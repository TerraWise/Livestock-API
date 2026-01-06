from openpyxl import Workbook


def vegetation_planting(
    beef_proportion: list[float] = [0],
    sheep_proportion: list[float] = [0],
    age: int = 0,
    area: float = 0,
    region: str = "South West",
    soil: str = "Loams & Clays",
    tree_species: str = "Mixed species (Environmental Plantings)",
) -> dict:
    return {
        "beefProportion": beef_proportion,
        "sheepProportion": sheep_proportion,
        "vegetation": {
            "age": age,
            "area": area,
            "region": region,
            "soil": soil,
            "treeSpecies": tree_species,
        },
    }


def extract_veg_data(inventory_sheet: Workbook) -> dict:
    ws = inventory_sheet["Vegetation"]
    veg_data = {"vegetation": []}

    row = 2
    while True:
        region = ws.cell(row, 2).value
        tree_species = ws.cell(row, 3).value
        soil = ws.cell(row, 5).value
        area = ws.cell(row, 6).value
        age = ws.cell(row, 8).value

        if any(
            list(
                map(
                    lambda x: x is None,
                    [
                        region,
                        tree_species,
                        soil,
                        area,
                        age,
                    ],
                )
            )
        ):
            if row == 2:
                veg_data["vegetation"].append(vegetation_planting())
            break

        veg_data["vegetation"].append(
            vegetation_planting(
                age=age,
                area=area,
                region=region,
                soil=soil,
                tree_species=tree_species,
            )
        )
        row += 1

    return veg_data
