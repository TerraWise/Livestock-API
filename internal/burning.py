from openpyxl import Workbook


def burning_record(
    fire_scar_area: float = 0,
    fuel: str = "coarse",
    patchines: str = "high",
    rainfall_zone: str = "low",
    season: str = "early dry season",
    vegetation: str = "Melaleuca woodland",
    years_since_last_fire: int = 0,
) -> dict:
    return {
        "fireScarArea": fire_scar_area,
        "fuel": fuel,
        "patchiness": patchines,
        "rainfallZone": rainfall_zone,
        "season": season,
        "vegetation": vegetation,
        "yearsSinceLastFire": years_since_last_fire,
    }


def extract_burning_data(inventory_sheet: Workbook) -> dict:
    ws = inventory_sheet["Burning"]
    burning_data = {"burning": []}

    row = 2
    while True:
        fuel = ws.cell(row, 1).value
        season = ws.cell(row, 2).value
        patchiness = ws.cell(row, 3).value
        rainfall_zone = ws.cell(row, 4).value
        years_since_last_fire = ws.cell(row, 5).value
        fire_scar_area = ws.cell(row, 6).value
        vegetation = ws.cell(row, 7).value

        if any(
            list(
                map(
                    lambda x: x is None,
                    [
                        fuel,
                        season,
                        patchiness,
                        rainfall_zone,
                        years_since_last_fire,
                        fire_scar_area,
                        vegetation,
                    ],
                )
            )
        ):
            if row == 2:
                burning_data["burning"].append(burning_record())
            break

        burning_data["burning"].append(
            burning_record(
                fire_scar_area=fire_scar_area,
                fuel=fuel,
                patchines=patchiness,
                rainfall_zone=rainfall_zone,
                season=season,
                vegetation=vegetation,
                years_since_last_fire=years_since_last_fire,
            )
        )
        row += 1

    return burning_data
