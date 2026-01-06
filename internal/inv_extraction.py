from copy import deepcopy
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook

from internal.sheep_vars import sheep_annual_stock_class_data
from internal.beef_vars import beef_annual_stock_class_data


def extract_inventories_from_excel(inventory_sheet: Workbook, livestock: str) -> list:
    seasonal_data = []

    seasonal_data.append(extract_seasonal_data(inventory_sheet, livestock))

    return seasonal_data


def extract_seasonal_data(inventory_sheet: Workbook, livestock: str) -> dict:
    seasonal_sheet = inventory_sheet[f"{livestock}SeasonalData"]
    stock_data = {}

    for col in range(3, 19):
        stock_class = seasonal_sheet.cell(2, col).value
        if livestock == livestock:
            stock_data[stock_class] = deepcopy(sheep_annual_stock_class_data)
        else:
            stock_data[stock_class] = deepcopy(beef_annual_stock_class_data)

        for row in range(3, 34):
            key = seasonal_sheet.cell(row, 1).value
            season = seasonal_sheet.cell(row, 2).value
            value = seasonal_sheet.cell(row, col).value

            if key in ["crudeProtein", "dryMatterDigestibility", "feedAvailability"]:
                if value == 0:
                    continue

            if season is not None:
                stock_data[stock_class][season.lower()][key] = value
            elif key in ["head", "purchaseWeight", "purchaseSource"]:
                stock_data[stock_class]["purchases"][0][key] = value
            elif key is not None:
                stock_data[stock_class][key] = value

    return stock_data


def extract_annual_data(
    inventory_sheet: Workbook, json_data: dict, livestock: str
) -> dict:
    if livestock.lower() == livestock:
        row = 2
    elif livestock.lower() == "cattle":
        row = 3
    else:
        raise ValueError("Unsupported livestock type")

    annual_sheet = inventory_sheet["Annual Data"]

    json_data = extract_lime_data(json_data, annual_sheet, row, livestock)
    json_data = extract_fertiliser_data(json_data, annual_sheet, row, livestock)
    json_data = extract_fuel_data(json_data, annual_sheet, row, livestock)
    json_data = extract_electricity_data(
        json_data, annual_sheet, inventory_sheet, row, livestock
    )
    json_data = extract_supplementation_data(json_data, annual_sheet, row, livestock)
    json_data = extract_feed_data(json_data, annual_sheet, row, livestock)
    json_data = extract_chemical_data(json_data, annual_sheet, row, livestock)
    json_data = extract_lambing_calving_rate(json_data, annual_sheet, row, livestock)
    if livestock == "sheep":
        json_data = extract_merino_pct(json_data, annual_sheet, row, livestock)
        json_data = extract_seasonalLambing_rate(
            json_data, annual_sheet, row, livestock
        )

    return json_data


def extract_lime_data(
    json_data: dict,
    annual_sheet: Worksheet,
    row: int,
    livestock: str,
    group: int = 0,
) -> dict:
    json_data[livestock][group]["limestone"] = annual_sheet.cell(row, 2).value
    json_data[livestock][group]["limestoneFraction"] = annual_sheet.cell(row, 3).value

    return json_data


def extract_fertiliser_data(
    json_data: dict,
    annual_sheet: Worksheet,
    row: int,
    livestock: str,
    group: int = 0,
) -> dict:
    json_data[livestock][group]["fertiliser"] = {
        "singleSuperphosphate": annual_sheet.cell(row, 4).value,
        "pastureDryland": annual_sheet.cell(row, 5).value,  # Urea
        "pastureIrrigated": 0,  # Urea
        "cropsDryland": annual_sheet.cell(row, 6).value,  # Urea
        "cropsIrrigated": 0,  # Urea
        "otherFertilisers": [],
    }

    for col in range(7, 21):
        json_data[livestock][group]["fertiliser"]["otherFertilisers"].append(
            {
                "otherType": annual_sheet.cell(1, col).value,
                "otherDryland": annual_sheet.cell(row, col).value,
                "otherIrrigated": 0,  # Assuming no irrigated data for other fertilisers
            }
        )

    return json_data


def extract_fuel_data(
    json_data: dict,
    annual_sheet: Worksheet,
    row: int,
    livestock: str,
    group: int = 0,
) -> dict:
    json_data[livestock][group]["diesel"] = annual_sheet.cell(row, 21).value

    json_data[livestock][group]["petrol"] = annual_sheet.cell(row, 22).value

    json_data[livestock][group]["lpg"] = annual_sheet.cell(row, 23).value

    return json_data


def extract_electricity_data(
    json_data: dict,
    annual_sheet: Worksheet,
    inventory_sheet: Workbook,
    row: int,
    livestock: str,
    group: int = 0,
) -> dict:
    json_data[livestock][group]["electricitySource"] = (
        inventory_sheet["Client detail"].cell(54, 7).value
    )

    if json_data[livestock][group]["electricitySource"] != "Renewable":
        json_data[livestock][group]["electricityRenewable"] = annual_sheet.cell(
            row, 30
        ).value

    json_data[livestock][group]["electricityUse"] = annual_sheet.cell(row, 31).value

    return json_data


def extract_supplementation_data(
    json_data: dict,
    annual_sheet: Worksheet,
    row: int,
    livestock: str,
    group: int = 0,
) -> dict:
    json_data[livestock][group]["mineralSupplementation"] = {
        "mineralBlock": annual_sheet.cell(row, 24).value,
        "mineralBlockUrea": annual_sheet.cell(row, 25).value,
        "weanerBlock": annual_sheet.cell(row, 26).value,
        "weanerBlockUrea": annual_sheet.cell(row, 27).value,
        "drySeasonMix": annual_sheet.cell(row, 28).value,
        "drySeasonMixUrea": annual_sheet.cell(row, 29).value,
    }

    return json_data


def extract_feed_data(
    json_data: dict,
    annual_sheet: Worksheet,
    row: int,
    livestock: str,
    group: int = 0,
) -> dict:
    json_data[livestock][group]["grainFeed"] = annual_sheet.cell(row, 32).value
    json_data[livestock][group]["hayFeed"] = annual_sheet.cell(row, 33).value
    if livestock == "beef":
        json_data[livestock][group]["cottonseedFeed"] = annual_sheet.cell(row, 34).value

    return json_data


def extract_chemical_data(
    json_data: dict,
    annual_sheet: Worksheet,
    row: int,
    livestock: str,
    group: int = 0,
) -> dict:
    json_data[livestock][group]["herbicide"] = annual_sheet.cell(row, 35).value
    json_data[livestock][group]["herbicideOther"] = annual_sheet.cell(row, 36).value

    return json_data


def extract_merino_pct(
    json_data: dict,
    annual_sheet: Worksheet,
    row: int,
    livestock: str,
    group: int = 0,
) -> dict:
    json_data[livestock][group]["merinoPercent"] = annual_sheet.cell(row, 37).value

    return json_data


def extract_lambing_calving_rate(
    json_data: dict,
    annual_sheet: Worksheet,
    row: int,
    livestock: str,
    group: int = 0,
) -> dict:
    season = annual_sheet.cell(row, 38).value
    rate = annual_sheet.cell(row, 39).value

    if livestock == "sheep":
        repro = "ewesLambing"
    else:
        repro = "cowsCalving"

    json_data[livestock][group][repro] = {
        "autumn": 0,
        "winter": 0,
        "spring": 0,
        "summer": 0,
    }
    json_data[livestock][group][repro][season.lower()] = rate  # type: ignore

    return json_data


def extract_seasonalLambing_rate(
    json_data: dict,
    annual_sheet: Worksheet,
    row: int,
    livestock: str,
    group: int = 0,
) -> dict:
    json_data[livestock][group]["seasonalLambing"] = {
        "autumn": 0,
        "winter": 0,
        "spring": 0,
        "summer": 0,
    }

    season = annual_sheet.cell(row, 40).value
    rate = annual_sheet.cell(row, 41).value

    json_data[livestock][group]["seasonalLambing"][season.lower()] = rate  # type: ignore

    return json_data
